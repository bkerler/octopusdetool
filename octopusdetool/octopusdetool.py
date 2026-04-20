#!/usr/bin/env python3
"""
Octopus Energy Germany Smart Meter Data Fetcher

Fetches electricity consumption data from Octopus Energy Germany API
and outputs to CSV format. Can also fill German electricity tariff Excel templates.
"""

import argparse
import csv
import ctypes
import importlib.resources as package_resources
import json
import os
import platform
import shutil
import sys
import time
import uuid
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo
from zoneinfo import ZoneInfoNotFoundError

import requests
import yaml

try:
    APP_TIMEZONE = ZoneInfo("Europe/Berlin")
    APP_TIMEZONE_HAS_IANA_DATA = True
except ZoneInfoNotFoundError:
    # Windows installations may not ship an IANA timezone database.
    # Fall back to the system local timezone for conversions, otherwise CET.
    APP_TIMEZONE = datetime.now().astimezone().tzinfo or timezone(timedelta(hours=1))
    APP_TIMEZONE_HAS_IANA_DATA = False

EXCEL_TEMPLATE_FILENAME = "smartmeter_daten.xlsx"
HEAT_EXCEL_TEMPLATE_FILENAME = "smartmeter_heat_daten.xlsx"
DEFAULT_TARIFF_GO_CT = 15.92
DEFAULT_TARIFF_STANDARD_CT = 29.13
DEFAULT_TARIFF_HEAT_LOW_CT = 21.50
DEFAULT_TARIFF_HEAT_STANDARD_CT = 28.51
DEFAULT_TARIFF_HEAT_HIGH_CT = 33.51
DEFAULT_MONTHLY_BASE_PRICE_EUR = 15.94
DEFAULT_TARIFF_HEAT_MONTHLY_BASE_PRICE_EUR = 14.50
TARIFF_INTELLIGENT_GO = "Intelligent Octopus Go"
TARIFF_INTELLIGENT_12 = "Intelligent Octopus 12"
TARIFF_INTELLIGENT_HEAT = "Octopus Heat"
TARIFF_DYNAMIC = "Dynamic Octopus"
TARIFF_TWO_ZONES = "two_zones"
TARIFF_THREE_ZONES = "three_zones"
DISPLAY_NAME_OCTOPUS_GO = "octopus go"
DISPLAY_NAME_OCTOPUS_HEAT = "octopus heat"
DISPLAY_NAME_DYNAMIC_OCTOPUS = "dynamicoctopus"
REQUEST_TIMEOUT_SECONDS = 30
REQUEST_TIMEOUT_RETRIES = 2
REQUEST_TIMEOUT_RETRY_DELAY_SECONDS = 2


class _WindowsGuid(ctypes.Structure):
    _fields_ = [
        ("Data1", ctypes.c_ulong),
        ("Data2", ctypes.c_ushort),
        ("Data3", ctypes.c_ushort),
        ("Data4", ctypes.c_ubyte * 8),
    ]

    @classmethod
    def from_string(cls, value: str) -> "_WindowsGuid":
        parsed = uuid.UUID(value.strip("{}"))
        data4 = (ctypes.c_ubyte * 8)(*parsed.bytes[8:])
        return cls(
            parsed.time_low,
            parsed.time_mid,
            parsed.time_hi_version,
            data4,
        )

@dataclass(slots=True)
class TariffAgreement:
    display_name: str
    valid_from: str
    valid_to: str | None
    agreement_id: str | None = None


@dataclass(slots=True)
class TariffSettings:
    tariff_type: str
    low_ct: float
    standard_ct: float
    high_ct: float
    monthly_base_price_eur: float


@dataclass(slots=True)
class TariffRate:
    name: str
    rate_ct: float
    windows: tuple[tuple[str, str], ...]


def get_demo_tariff_profile(mode: str) -> tuple[str, TariffSettings, list[TariffRate]]:
    normalized = str(mode).strip().lower()
    if normalized in {"2", "two", "2-zones", "two_zones"}:
        settings = TariffSettings(
            tariff_type=TARIFF_TWO_ZONES,
            low_ct=12.0,
            standard_ct=29.0,
            high_ct=0.0,
            monthly_base_price_eur=14.90,
        )
        rates = [
            TariffRate("Demo Low", 12.0, (("00:00", "05:00"),)),
            TariffRate("Demo Standard", 29.0, (("05:00", "24:00"),)),
        ]
        return TARIFF_INTELLIGENT_12, settings, rates

    if normalized in {"3", "three", "3-zones", "three_zones"}:
        settings = TariffSettings(
            tariff_type=TARIFF_THREE_ZONES,
            low_ct=21.5,
            standard_ct=28.5,
            high_ct=33.5,
            monthly_base_price_eur=14.50,
        )
        rates = [
            TariffRate("Demo Low", 21.5, (("02:00", "06:00"), ("12:00", "16:00"))),
            TariffRate("Demo Standard", 28.5, (("06:00", "18:00"), ("21:00", "02:00"))),
            TariffRate("Demo High", 33.5, (("18:00", "21:00"),)),
        ]
        return TARIFF_INTELLIGENT_HEAT, settings, rates

    if normalized in {"dynamically", "dynamic", "dynamicoctopus"}:
        settings = TariffSettings(
            tariff_type=TARIFF_DYNAMIC,
            low_ct=10.0,
            standard_ct=18.0,
            high_ct=26.0,
            monthly_base_price_eur=15.00,
        )
        rates = [
            TariffRate("Demo Offpeak", 10.0, (("00:00", "04:00"),)),
            TariffRate("Demo Shoulder", 18.0, (("04:00", "08:00"), ("22:00", "24:00"))),
            TariffRate("Demo Peak", 26.0, (("08:00", "18:00"),)),
            TariffRate("Demo Superpeak", 31.0, (("18:00", "22:00"),)),
        ]
        return TARIFF_DYNAMIC, settings, rates

    raise ValueError(f"Unknown demo tariff mode: {mode}")


def get_default_tariff_settings_for_type(tariff_type: str) -> TariffSettings:
    if tariff_type == TARIFF_THREE_ZONES:
        return TariffSettings(
            tariff_type=TARIFF_THREE_ZONES,
            low_ct=DEFAULT_TARIFF_HEAT_LOW_CT,
            standard_ct=DEFAULT_TARIFF_HEAT_STANDARD_CT,
            high_ct=DEFAULT_TARIFF_HEAT_HIGH_CT,
            monthly_base_price_eur=DEFAULT_TARIFF_HEAT_MONTHLY_BASE_PRICE_EUR,
        )

    return TariffSettings(
        tariff_type=TARIFF_TWO_ZONES,
        low_ct=DEFAULT_TARIFF_GO_CT,
        standard_ct=DEFAULT_TARIFF_STANDARD_CT,
        high_ct=0.0,
        monthly_base_price_eur=DEFAULT_MONTHLY_BASE_PRICE_EUR,
    )


def get_bundled_excel_template_path(tariff_type: str = TARIFF_INTELLIGENT_GO) -> Path:
    """Get the bundled blank Excel template path."""
    filename = (
        HEAT_EXCEL_TEMPLATE_FILENAME
        if tariff_type == TARIFF_INTELLIGENT_HEAT
        else EXCEL_TEMPLATE_FILENAME
    )
    return Path(__file__).parent / filename


def _get_bundled_excel_template_filename(tariff_type: str) -> str:
    return (
        HEAT_EXCEL_TEMPLATE_FILENAME
        if tariff_type == TARIFF_INTELLIGENT_HEAT
        else EXCEL_TEMPLATE_FILENAME
    )


def _get_bundled_excel_template_resource(tariff_type: str):
    return package_resources.files("octopusdetool").joinpath(
        _get_bundled_excel_template_filename(tariff_type)
    )


def get_app_data_folder() -> Path:
    """Get the OS-specific application data/config folder."""
    system = platform.system()

    if system == "Windows":
        # FOLDERID_RoamingAppData
        app_data = _get_windows_known_folder(
            "{3EB685DB-65F9-4CF6-A03A-E3EF65729F3D}"
        )
        if app_data:
            return app_data
        app_data = os.environ.get("APPDATA")
        if app_data:
            return Path(app_data)
        userprofile = os.environ.get("USERPROFILE")
        if userprofile:
            return Path(userprofile) / "AppData" / "Roaming"
        home = Path.home()
        if home.name:
            return home / "AppData" / "Roaming"
        return Path.cwd()
    elif system == "Darwin":
        return Path.home() / "Library" / "Application Support"
    else:
        # Linux/Unix: prefer XDG_CONFIG_HOME, fallback to ~/.config
        xdg_config = os.environ.get("XDG_CONFIG_HOME")
        if xdg_config:
            return Path(xdg_config)
        return Path.home() / ".config"


def get_documents_folder() -> Path:
    """Get the user's Documents folder path (cross-platform)."""
    system = platform.system()

    if system == "Windows":
        # Prefer the Windows "Documents" known folder because Path.home()
        # can point to a stale or virtualized profile in packaged apps.
        docs = _get_windows_documents_folder()
    elif system == "Darwin":
        # macOS: ~/Documents
        docs = Path.home() / "Documents"
    else:
        # Linux/Unix: try XDG_DOCUMENTS_DIR, fallback to ~/Documents
        xdg_docs = os.environ.get("XDG_DOCUMENTS_DIR")
        if xdg_docs:
            docs = Path(xdg_docs)
        else:
            docs = Path.home() / "Documents"

    return docs


def _get_windows_documents_folder() -> Path:
    """Resolve the current user's Documents folder on Windows."""
    known_folder = _get_windows_known_folder(
        "{FDD39AD0-238F-46AF-ADB4-6C85480369C7}"
    )
    if known_folder:
        return known_folder

    userprofile = os.environ.get("USERPROFILE")
    if userprofile:
        return Path(userprofile) / "Documents"

    homedrive = os.environ.get("HOMEDRIVE")
    homepath = os.environ.get("HOMEPATH")
    if homedrive and homepath:
        return Path(f"{homedrive}{homepath}") / "Documents"

    home = Path.home()
    if home.name:
        return home / "Documents"

    # Final fallback keeps the app writable even if Windows profile discovery fails.
    return Path.cwd() / "Documents"


def _get_windows_known_folder(folder_id: str) -> Path | None:
    """Return a Windows known folder path, or None if it cannot be resolved."""
    path_ptr = ctypes.c_wchar_p()
    shell32 = getattr(ctypes.windll, "shell32", None)
    ole32 = getattr(ctypes.windll, "ole32", None)
    if shell32 is None or ole32 is None:
        return None

    try:
        result = shell32.SHGetKnownFolderPath(
            ctypes.byref(_WindowsGuid.from_string(folder_id)),
            0,
            None,
            ctypes.byref(path_ptr),
        )
        if result != 0 or not path_ptr.value:
            return None
        return Path(path_ptr.value)
    except (AttributeError, OSError, TypeError, ValueError):
        return None
    finally:
        if path_ptr:
            ole32.CoTaskMemFree(path_ptr)


def get_app_config_folder() -> Path:
    """Get the application configuration folder path (OS-specific app data location)."""
    return _get_preferred_directory_path(get_app_data_folder() / "octopusdetool")


def get_smartmeter_data_folder() -> Path:
    """Get the smartmeter_data folder path (in Documents)."""
    return _get_preferred_directory_path(get_documents_folder() / "smartmeter_data")


def ensure_smartmeter_data_folder() -> Path:
    """Create and return the smartmeter data directory used by the app."""
    folder = get_smartmeter_data_folder()
    if not folder.exists():
        try:
            os.makedirs(folder, exist_ok=True)
        except Exception:
            pass
    return folder


def init_app_config_folder() -> tuple[bool, str | None]:
    """Initialize the app config folder once at program start.

    Returns:
        Tuple of (success: bool, error_message: str | None)
    """
    folder = get_app_config_folder()
    if folder.exists() and folder.is_dir():
        return True, None
    try:
        os.makedirs(folder, exist_ok=True)
        if folder.exists() and folder.is_dir():
            return True, None
        return False, f"Der Ordner konnte nicht erstellt werden: {folder}"
    except Exception as exc:
        return False, f"Der Ordner konnte nicht erstellt werden: {folder}\n{exc}"


def _get_preferred_directory_path(path: Path) -> Path:
    """Return a usable directory path, even if the preferred path is a file."""
    if not path.exists() or path.is_dir():
        return path

    for suffix in ("_dir", "_folder", "_data"):
        candidate = path.with_name(f"{path.name}{suffix}")
        if not candidate.exists() or candidate.is_dir():
            return candidate

    index = 1
    while True:
        candidate = path.with_name(f"{path.name}_{index}")
        if not candidate.exists() or candidate.is_dir():
            return candidate
        index += 1


def _copy_cell_style(source_cell, target_cell) -> None:
    target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format


def detect_excel_template_type(workbook_path: Path | str) -> str:
    """Detect the tariff model from the number of tariff zones listed in Einstellungen."""
    try:
        import openpyxl
    except ImportError:
        return TARIFF_TWO_ZONES

    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        ws = workbook["Einstellungen"]
        tariff_zone_count = 0
        for row in range(3, 10):
            label = ws[f"A{row}"].value
            if isinstance(label, str) and label.startswith("Tarif "):
                tariff_zone_count += 1
                continue
            if label == "Startdatum":
                break
        workbook.close()
        return TARIFF_THREE_ZONES if tariff_zone_count >= 3 else TARIFF_TWO_ZONES
    except Exception:
        return TARIFF_TWO_ZONES


def get_excel_layout(tariff_type: str) -> dict[str, str]:
    if tariff_type == TARIFF_THREE_ZONES:
        return {
            "tariff_low": "B3",
            "tariff_standard": "B4",
            "tariff_high": "B5",
            "start_date": "B6",
            "end_date": "B7",
            "base_price": "B8",
            "zone_count": "3",
        }

    return {
        "tariff_low": "B3",
        "tariff_standard": "B4",
        "tariff_high": "",
        "start_date": "B5",
        "end_date": "B6",
        "base_price": "B7",
        "zone_count": "2",
    }


def create_heat_excel_template(source_path: Path, target_path: Path) -> Path:
    """Create a heat-specific workbook copy from the stock template."""
    import openpyxl

    source_path = Path(source_path)
    target_path = Path(target_path)
    try:
        target_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass
    workbook = openpyxl.load_workbook(source_path)

    ws_settings = workbook["Einstellungen"]
    ws_consumption = workbook["Verbrauch"]
    ws_day = workbook["Tagesübersicht"]
    ws_week = workbook["Wochenübersicht"]
    ws_month = workbook["Monatsübersicht"]
    ws_year = workbook["Jahresübersicht"]

    # Einstellungen sheet: expand from 2 to 3 tariff rows without shifting the whole workbook.
    heat_settings = {
        "A3": "Tarif Niedrig 02:00-05:59, 12:00-15:59 (ct/kWh)",
        "B3": DEFAULT_TARIFF_HEAT_LOW_CT,
        "A4": "Tarif Standard 06:00-11:59, 16:00-17:59, 21:00-01:59 (ct/kWh)",
        "B4": DEFAULT_TARIFF_HEAT_STANDARD_CT,
        "A5": "Tarif Hoch 18:00-20:59 (ct/kWh)",
        "B5": DEFAULT_TARIFF_HEAT_HIGH_CT,
        "A6": "Startdatum",
        "B6": ws_settings["B5"].value,
        "A7": "Enddatum der Vorlage",
        "B7": ws_settings["B6"].value,
        "A8": "Grundpreis pro Monat (€)",
        "B8": DEFAULT_TARIFF_HEAT_MONTHLY_BASE_PRICE_EUR,
        "A9": "Tipp",
        "B9": "Wenn sich Tarif oder Grundpreis ändern, nur B3/B4/B5/B8 anpassen.",
    }

    for target_coord, value in sorted(
        heat_settings.items(),
        key=lambda item: int("".join(character for character in item[0] if character.isdigit())),
        reverse=True,
    ):
        if target_coord in {"A5", "B5"}:
            style_source = ws_settings["A4"] if target_coord.startswith("A") else ws_settings["B4"]
        elif target_coord in {"A6", "B6"}:
            style_source = ws_settings["A5"] if target_coord.startswith("A") else ws_settings["B5"]
        elif target_coord in {"A7", "B7"}:
            style_source = ws_settings["A6"] if target_coord.startswith("A") else ws_settings["B6"]
        elif target_coord in {"A8", "B8"}:
            style_source = ws_settings["A7"] if target_coord.startswith("A") else ws_settings["B7"]
        elif target_coord in {"A9", "B9"}:
            style_source = ws_settings["A8"] if target_coord.startswith("A") else ws_settings["B8"]
        else:
            style_source = ws_settings[target_coord]
        _copy_cell_style(style_source, ws_settings[target_coord])
        ws_settings[target_coord].value = value

    # Verbrauch formulas switch to the heat zones and the shifted start date / base price cells.
    max_consumption_row = ws_consumption.max_row
    for row in range(9, max_consumption_row + 1):
        hour_ref = f"B{row}"
        ws_consumption[f"A{row}"] = f"=Einstellungen!$B$6+INT((ROW()-9)/24)"
        ws_consumption[f"D{row}"] = (
            f'=IF(OR(AND({hour_ref}>=2,{hour_ref}<6),AND({hour_ref}>=12,{hour_ref}<16)),'
            'Einstellungen!$B$3,'
            f'IF(AND({hour_ref}>=18,{hour_ref}<21),Einstellungen!$B$5,Einstellungen!$B$4))'
        )
        ws_consumption[f"F{row}"] = (
            f'=IF(OR(AND({hour_ref}>=2,{hour_ref}<6),AND({hour_ref}>=12,{hour_ref}<16)),"Niedrig",'
            f'IF(AND({hour_ref}>=18,{hour_ref}<21),"Hoch","Standard"))'
        )

    # Overview sheets: keep existing totals and costs, add Hoch as an extra column.
    ws_day["D4"] = "davon Niedrig (kWh)"
    ws_day["E4"] = "davon Standard (kWh)"
    ws_day["I4"] = "davon Hoch (kWh)"
    for row in range(5, ws_day.max_row + 1):
        ws_day[f"D{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&A{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&B{row},Verbrauch!$F$9:$F$18896,"Niedrig")'
        )
        ws_day[f"E{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&A{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&B{row},Verbrauch!$F$9:$F$18896,"Standard")'
        )
        ws_day[f"I{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&A{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&B{row},Verbrauch!$F$9:$F$18896,"Hoch")'
        )

    ws_week["F4"] = "davon Niedrig (kWh)"
    ws_week["G4"] = "davon Standard (kWh)"
    ws_week["J4"] = "davon Hoch (kWh)"
    for row in range(5, ws_week.max_row + 1):
        ws_week[f"F{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&A{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&B{row},Verbrauch!$F$9:$F$18896,"Niedrig")'
        )
        ws_week[f"G{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&A{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&B{row},Verbrauch!$F$9:$F$18896,"Standard")'
        )
        ws_week[f"J{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&A{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&B{row},Verbrauch!$F$9:$F$18896,"Hoch")'
        )

    ws_month["F1"] = (
        'Diese Übersicht aktualisiert sich automatisch, sobald du in der Tabelle "Verbrauch" '
        'deine stündlichen kWh einträgst. Monatskosten enthalten jetzt den Grundpreis von Einstellungen!B8.'
    )

    ws_year["E4"] = "davon Niedrig (kWh)"
    ws_year["F4"] = "davon Standard (kWh)"
    ws_year["J4"] = "davon Hoch (kWh)"
    for row in range(5, ws_year.max_row + 1):
        ws_year[f"E{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&B{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&C{row},Verbrauch!$F$9:$F$18896,"Niedrig")'
        )
        ws_year[f"F{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&B{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&C{row},Verbrauch!$F$9:$F$18896,"Standard")'
        )
        ws_year[f"J{row}"] = (
            f'=SUMIFS(Verbrauch!$C$9:$C$18896,Verbrauch!$A$9:$A$18896,">="&B{row},'
            f'Verbrauch!$A$9:$A$18896,"<"&C{row},Verbrauch!$F$9:$F$18896,"Hoch")'
        )

    workbook.save(target_path)
    workbook.close()
    return target_path


def ensure_excel_template(tariff_type: str = TARIFF_TWO_ZONES):
    """Copy the requested Excel template to smartmeter_data if it doesn't exist."""
    smartmeter_folder = ensure_smartmeter_data_folder()

    source_name = _get_bundled_excel_template_filename(tariff_type)
    target = smartmeter_folder / source_name

    if not target.exists():
        resource = _get_bundled_excel_template_resource(tariff_type)
        if resource.is_file():
            try:
                with package_resources.as_file(resource) as source_path:
                    shutil.copy2(source_path, target)
                print(f"Excel-Vorlage kopiert nach: {target}")
            except OSError:
                pass
        elif tariff_type == TARIFF_THREE_ZONES:
            stock_resource = _get_bundled_excel_template_resource(TARIFF_TWO_ZONES)
            if stock_resource.is_file():
                try:
                    with package_resources.as_file(stock_resource) as stock_source_path:
                        create_heat_excel_template(stock_source_path, target)
                    print(f"Heat-Excel-Vorlage erzeugt nach: {target}")
                except OSError:
                    pass

    if target.exists():
        return target

    if tariff_type == TARIFF_THREE_ZONES:
        stock_template = get_bundled_excel_template_path(TARIFF_TWO_ZONES)
        if stock_template.exists():
            try:
                create_heat_excel_template(stock_template, target)
                print(f"Heat-Excel-Vorlage erzeugt nach: {target}")
                return target
            except OSError:
                pass

    source = get_bundled_excel_template_path(tariff_type)
    if source.exists():
        return source

    return target


def get_default_output_path() -> Path:
    """Get the default internal readings cache path."""
    return get_app_config_folder() / "readings.json"


def get_default_consumption_csv_path() -> Path:
    """Get the default API-format CSV cache path."""
    return get_app_config_folder() / "consumption.csv"


def cleanup_app_config_folder() -> None:
    """Keep only config.json, readings.json, and consumption.csv in the app config folder."""
    folder = get_app_config_folder()
    allowed = {"config.json", "readings.json", "consumption.csv"}
    try:
        folder.mkdir(parents=True, exist_ok=True)
    except OSError:
        return

    for child in folder.iterdir():
        if child.name in allowed:
            continue
        if child.is_file():
            try:
                child.unlink()
            except OSError:
                pass


def get_default_excel_path(tariff_type: str = TARIFF_TWO_ZONES) -> Path:
    """Get the default Excel output path shown in the UI/CLI."""
    filename = (
        HEAT_EXCEL_TEMPLATE_FILENAME
        if tariff_type == TARIFF_THREE_ZONES
        else EXCEL_TEMPLATE_FILENAME
    )
    return get_smartmeter_data_folder() / filename


def load_excel_tariff_settings(template_path: Path | None = None) -> dict[str, float | str]:
    """Read tariff defaults from the Excel template, falling back to built-in values."""
    defaults = {
        "tariff_type": TARIFF_TWO_ZONES,
        "tariff_go_ct": DEFAULT_TARIFF_GO_CT,
        "tariff_standard_ct": DEFAULT_TARIFF_STANDARD_CT,
        "tariff_heat_low_ct": DEFAULT_TARIFF_HEAT_LOW_CT,
        "tariff_heat_standard_ct": DEFAULT_TARIFF_HEAT_STANDARD_CT,
        "tariff_heat_high_ct": DEFAULT_TARIFF_HEAT_HIGH_CT,
        "monthly_base_price_eur": DEFAULT_MONTHLY_BASE_PRICE_EUR,
    }

    try:
        import openpyxl
    except ImportError:
        return defaults

    workbook_path = template_path or ensure_excel_template()
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        return defaults

    try:
        detected_type = detect_excel_template_type(workbook_path)
        defaults["tariff_type"] = detected_type
        if detected_type == TARIFF_THREE_ZONES:
            defaults["monthly_base_price_eur"] = DEFAULT_TARIFF_HEAT_MONTHLY_BASE_PRICE_EUR
        layout = get_excel_layout(detected_type)
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        ws = wb["Einstellungen"]
        values = {
            "tariff_go_ct": ws[layout["tariff_low"]].value if layout["tariff_low"] else None,
            "tariff_standard_ct": ws[layout["tariff_standard"]].value if layout["tariff_standard"] else None,
            "tariff_heat_low_ct": ws[layout["tariff_low"]].value if detected_type == TARIFF_THREE_ZONES else None,
            "tariff_heat_standard_ct": ws[layout["tariff_standard"]].value if detected_type == TARIFF_THREE_ZONES else None,
            "tariff_heat_high_ct": ws[layout["tariff_high"]].value if layout["tariff_high"] else None,
            "monthly_base_price_eur": ws[layout["base_price"]].value,
        }
        wb.close()
    except Exception:
        return defaults

    for key, default in defaults.items():
        if key == "tariff_type":
            continue
        try:
            if values[key] is not None:
                defaults[key] = float(values[key])
        except (TypeError, ValueError):
            defaults[key] = default

    return defaults


def get_tariff_rate_ct(
    reading_start: datetime,
    tariff_go_ct: float = DEFAULT_TARIFF_GO_CT,
    tariff_standard_ct: float = DEFAULT_TARIFF_STANDARD_CT,
    tariff_type: str = TARIFF_INTELLIGENT_GO,
    tariff_high_ct: float = 0.0,
) -> float:
    """Return the tariff in ct/kWh for the given interval start."""
    local_start = to_local_datetime(reading_start)
    hour = local_start.hour

    if tariff_type == TARIFF_THREE_ZONES:
        if hour in {2, 3, 4, 5, 12, 13, 14, 15}:
            return tariff_go_ct
        if hour in {18, 19, 20}:
            return tariff_high_ct
        return tariff_standard_ct

    return tariff_go_ct if 0 <= hour <= 5 else tariff_standard_ct


def classify_tariff_zone(reading_start: datetime, tariff_type: str) -> str:
    local_start = to_local_datetime(reading_start)
    hour = local_start.hour

    if tariff_type == TARIFF_THREE_ZONES:
        if hour in {2, 3, 4, 5, 12, 13, 14, 15}:
            return "low"
        if hour in {18, 19, 20}:
            return "high"
        return "standard"

    return "low" if 0 <= hour <= 5 else "standard"


def _normalize_rate_windows(windows: list[dict] | None) -> tuple[tuple[str, str], ...]:
    normalized: list[tuple[str, str]] = []
    for window in windows or []:
        active_from = window.get("activeFromTime")
        active_to = window.get("activeToTime")
        if not active_from or not active_to:
            continue
        normalized.append((str(active_from)[:5], str(active_to)[:5]))
    normalized.sort()
    return tuple(normalized)


def _extract_tariff_rates(unit_rate_information: dict | None) -> list[TariffRate]:
    if not isinstance(unit_rate_information, dict):
        return []

    typename = unit_rate_information.get("__typename")
    if typename != "TimeOfUseProductUnitRateInformation":
        return []

    rates: list[TariffRate] = []
    for rate in unit_rate_information.get("rates", []) or []:
        try:
            rate_ct = float(rate["latestGrossUnitRateCentsPerKwh"])
        except (KeyError, TypeError, ValueError):
            continue

        rates.append(
            TariffRate(
                name=rate.get("timeslotName") or "",
                rate_ct=rate_ct,
                windows=_normalize_rate_windows(rate.get("timeslotActivationRules")),
            )
        )

    return rates


def _tariff_type_from_rates(rates: list[TariffRate]) -> str | None:
    if not rates:
        return None
    return TARIFF_INTELLIGENT_HEAT if len(rates) >= 3 else TARIFF_INTELLIGENT_GO


def _extract_monthly_base_price(agreement_data: dict) -> float | None:
    """Extract monthly base price (Grundpreis) from standing charge data.

    The API returns standing charge information which may be a single object
    or a list of historical rates. We convert the daily rate to a monthly
    amount by multiplying with the average number of days per month (365/12).
    """
    raw = agreement_data.get("standingChargeGrossRateInformation")
    if raw is None:
        return None

    if isinstance(raw, dict):
        entries = [raw]
    elif isinstance(raw, list):
        entries = raw
    else:
        return None

    valid_entries = [
        e for e in entries if isinstance(e, dict) and e.get("grossRate") is not None
    ]
    if not valid_entries:
        return None

    # Prefer entries without an end date (currently valid), then latest start date
    valid_entries.sort(
        key=lambda e: (bool(e.get("rateValidToDate")), e.get("date", "")),
        reverse=True,
    )
    best_entry = valid_entries[0]

    try:
        daily_rate = float(best_entry["grossRate"])
    except (TypeError, ValueError):
        return None

    if daily_rate <= 0:
        return None

    # The API may return the standing charge in different units depending on
    # the market. We compute all three plausible monthly interpretations and
    # pick the one that falls into a reasonable German residential base-price
    # range (roughly 4-50 EUR/month).
    as_monthly = daily_rate
    as_euros_per_day = daily_rate * (365.0 / 12.0)
    as_cents_per_day = daily_rate * (365.0 / 12.0) / 100.0

    candidates = [
        (as_monthly, "monthly"),
        (as_euros_per_day, "euros/day"),
        (as_cents_per_day, "cents/day"),
    ]

    # Prefer candidates in the reasonable range [4, 50]
    reasonable = [(v, u) for v, u in candidates if 4.0 <= v <= 50.0]
    if reasonable:
        # Pick the one closest to a typical base price of 15 EUR/month
        best_value, best_unit = min(reasonable, key=lambda x: abs(x[0] - 15.0))
        print(f"[DEBUG] Extracted standing charge: raw={daily_rate}, unit={best_unit}, monthly={round(best_value, 2)}")
        return round(best_value, 2)

    # Fallback: choose the candidate closest to the reasonable range
    def _distance(v: float) -> float:
        if v < 4.0:
            return 4.0 - v
        return v - 50.0

    best_value, best_unit = min(candidates, key=lambda x: _distance(x[0]))
    print(f"[DEBUG] Extracted standing charge: raw={daily_rate}, unit={best_unit}, monthly={round(best_value, 2)}")
    return round(best_value, 2)


def map_rate_structure_to_tariff_settings(
    agreement_display_name: str,
    unit_rate_information: dict | None,
    monthly_base_price_eur: float,
) -> TariffSettings | None:
    """Map API time-of-use rates to the existing 2-zone/3-zone Excel tariff model."""
    rates = _extract_tariff_rates(unit_rate_information)
    if not rates:
        return None

    inferred_type = _tariff_type_from_rates(rates)
    if inferred_type == TARIFF_THREE_ZONES and len(rates) >= 3:
        unique_prices = sorted({rate.rate_ct for rate in rates})
        if len(unique_prices) == 3:
            return TariffSettings(
                tariff_type=TARIFF_THREE_ZONES,
                low_ct=unique_prices[0],
                standard_ct=unique_prices[1],
                high_ct=unique_prices[2],
                monthly_base_price_eur=monthly_base_price_eur,
            )
        return TariffSettings(
            tariff_type=TARIFF_THREE_ZONES,
            low_ct=rates[0].rate_ct,
            standard_ct=rates[1].rate_ct,
            high_ct=rates[2].rate_ct,
            monthly_base_price_eur=monthly_base_price_eur,
        )

    if inferred_type == TARIFF_TWO_ZONES and len(rates) >= 2:
        unique_prices = sorted({rate.rate_ct for rate in rates})
        low_price = unique_prices[0]
        standard_price = unique_prices[1] if len(unique_prices) > 1 else unique_prices[0]
        return TariffSettings(
            tariff_type=TARIFF_TWO_ZONES,
            low_ct=low_price,
            standard_ct=standard_price,
            high_ct=0.0,
            monthly_base_price_eur=monthly_base_price_eur,
        )

    return None


# German Octopus Energy API endpoints
GRAPHQL_URL = "https://api.oeg-kraken.energy/v1/graphql/"

# Authentication mutation
AUTH_MUTATION = """
mutation krakenTokenAuthentication($email: String!, $password: String!) {
    obtainKrakenToken(input: { email: $email, password: $password }) {
        token
        payload
    }
}
"""

# Viewer query - get current user's account without needing account number
VIEWER_QUERY = """
query Viewer {
    viewer {
        accounts {
            ... on AccountType {
                id
                number
            }
        }
    }
}
"""

# Meter discovery query
ACCOUNT_DETAILS_QUERY = """
query OverviewPage($accountNumber: String!) {
    account(accountNumber: $accountNumber) {
        id
        allProperties {
            id
            electricityMalos {
                maloNumber
                meter { id number shouldReceiveSmartMeterData }
                agreements {
                    id
                    validFrom
                    validTo
                    isActive
                    product {
                        displayName
                    }
                }
            }
        }
    }
}
"""

RATE_STRUCTURE_QUERY = """
query GetRateStructureForProductAgreement($agreementId: ID!) {
    agreement(id: $agreementId) {
        standingChargeGrossRateInformation {
            grossRate
            date
            rateValidToDate
            vatRate
        }
        unitRateInformation {
            __typename
            ... on UnitRateInformation {
                __typename
                ... on SimpleProductUnitRateInformation {
                    latestGrossUnitRateCentsPerKwh
                }
                ... on TimeOfUseProductUnitRateInformation {
                    rates {
                        timeslotName
                        latestGrossUnitRateCentsPerKwh
                        timeslotActivationRules {
                            activeFromTime
                            activeToTime
                        }
                    }
                }
            }
        }
    }
}
"""

# Consumption query - using measurements with hourly interval filter and pagination
MEASUREMENTS_QUERY = """
query getAccountMeasurements(
    $propertyId: ID!
    $first: Int!
    $after: String
    $utilityFilters: [UtilityFiltersInput!]
    $startOn: Date
    $endOn: Date
    $startAt: DateTime
    $endAt: DateTime
    $timezone: String
) {
    property(id: $propertyId) {
        measurements(
            first: $first
            after: $after
            utilityFilters: $utilityFilters
            startOn: $startOn
            endOn: $endOn
            startAt: $startAt
            endAt: $endAt
            timezone: $timezone
        ) {
            edges {
                cursor
                node {
                    value
                    unit
                    ... on IntervalMeasurementType {
                        startAt
                        endAt
                        durationInSeconds
                    }
                    metaData {
                        statistics {
                            costExclTax {
                                pricePerUnit {
                                    amount
                                }
                                costCurrency
                                estimatedAmount
                            }
                            costInclTax {
                                costCurrency
                                estimatedAmount
                            }
                            value
                            description
                            label
                            type
                        }
                    }
                }
            }
            pageInfo {
                hasNextPage
                endCursor
            }
        }
    }
}
"""

SMART_USAGE_QUERY = """
query GetSmartUsage(
    $propertyId: ID!
    $timezone: String!
    $startAt: DateTime!
    $endAt: DateTime!
    $utilityFilters: [UtilityFiltersInput!]!
) {
    property(id: $propertyId) {
        measurements(
            first: 1000
            timezone: $timezone
            startAt: $startAt
            endAt: $endAt
            utilityFilters: $utilityFilters
        ) {
            edges {
                node {
                    value
                    unit
                    source
                    ... on IntervalMeasurementType {
                        startAt
                        endAt
                    }
                    metaData {
                        utilityFilters {
                            ... on ElectricityFiltersOutput {
                                readingDirection
                            }
                        }
                        statistics {
                            label
                            value
                            type
                            costInclTax {
                                costCurrency
                                estimatedAmount
                            }
                            costExclTax {
                                costCurrency
                                estimatedAmount
                            }
                        }
                    }
                }
            }
        }
    }
}
"""


def _normalize_reading_direction(direction: str | None) -> str:
    normalized = str(direction or "CONSUMPTION").upper()
    return normalized if normalized in {"CONSUMPTION", "GENERATION"} else "CONSUMPTION"


def _extract_reading_direction(node: dict) -> str:
    meta_data = node.get("metaData") or {}
    utility_filters = meta_data.get("utilityFilters")

    if isinstance(utility_filters, list):
        candidate = utility_filters[0] if utility_filters else {}
    elif isinstance(utility_filters, dict):
        candidate = utility_filters
    else:
        candidate = {}

    if not isinstance(candidate, dict):
        candidate = {}

    electricity_filters = candidate.get("electricityFilters")
    if not isinstance(electricity_filters, dict):
        electricity_filters = {}

    return _normalize_reading_direction(electricity_filters.get("readingDirection"))

METER_READINGS_QUERY = """
query getMeterReadingsElectricity($accountNumber: String!, $meterId: ID!, $cursor: String) {
    electricityMeterReadings(
        first: 20
        after: $cursor
        accountNumber: $accountNumber
        meterId: $meterId
    ) {
        edges {
            node {
                readAt
                value
                registerObisCode
                typeOfRead
                registerType
                origin
                meterId
            }
        }
        pageInfo {
            endCursor
            hasNextPage
        }
    }
}
"""


class OctopusGermanyClient:
    """Client for Octopus Energy Germany GraphQL API."""

    def __init__(self, email: str, password: str, debug: bool = False):
        self.email = email
        self.password = password
        self.token = None
        self.debug = debug
        self.last_error_kind: str | None = None
        self.last_error_message: str | None = None

    def _log_debug(self, message: str):
        """Print debug message if debug mode is enabled."""
        if self.debug:
            print(f"[DEBUG] {message}")

    def _clear_last_error(self) -> None:
        self.last_error_kind = None
        self.last_error_message = None

    def _set_last_error(self, kind: str, message: str) -> None:
        self.last_error_kind = kind
        self.last_error_message = message

    def _post_with_retry(self, *, json_payload: dict, headers: dict | None = None):
        last_retryable_error: requests.exceptions.RequestException | None = None
        attempts = REQUEST_TIMEOUT_RETRIES + 1

        for attempt in range(1, attempts + 1):
            try:
                return requests.post(
                    GRAPHQL_URL,
                    json=json_payload,
                    headers=headers,
                    timeout=REQUEST_TIMEOUT_SECONDS,
                )
            except (
                requests.exceptions.Timeout,
                requests.exceptions.ConnectionError,
            ) as exc:
                last_retryable_error = exc
                if attempt >= attempts:
                    raise
                error_name = exc.__class__.__name__
                print(
                    f"{error_name} beim Request, neuer Versuch "
                    f"{attempt + 1}/{attempts} in {REQUEST_TIMEOUT_RETRY_DELAY_SECONDS}s..."
                )
                self._log_debug(
                    f"Retrying after {error_name.lower()} "
                    f"({attempt}/{attempts - 1} retries used): {exc}"
                )
                time.sleep(REQUEST_TIMEOUT_RETRY_DELAY_SECONDS)

        if last_retryable_error is not None:
            raise last_retryable_error

    def authenticate(self) -> bool:
        """Authenticate and get JWT token."""
        self._clear_last_error()
        variables = {
            "email": self.email,
            "password": self.password
        }
        
        payload = {
            "query": AUTH_MUTATION,
            "variables": variables
        }
        
        if self.debug:
            print("\n" + "="*80)
            print("AUTH REQUEST:")
            print("="*80)
            print(f"URL: {GRAPHQL_URL}")
            print(f"Payload: {json.dumps(payload, indent=2)}")
            print("="*80)
        
        try:
            response = self._post_with_retry(json_payload=payload)
            
            if self.debug:
                print("\n" + "="*80)
                print("AUTH RESPONSE:")
                print("="*80)
                print(f"Status: {response.status_code}")
                try:
                    print(f"Body: {json.dumps(response.json(), indent=2)}")
                except:
                    print(f"Body: {response.text}")
                print("="*80 + "\n")
            
            response.raise_for_status()
            data = response.json()
            
            if "errors" in data:
                error_message = f"Authentifizierungsfehler: {data['errors']}"
                self._set_last_error("auth", error_message)
                print(error_message)
                return False
            
            self.token = data["data"]["obtainKrakenToken"]["token"]
            self._clear_last_error()
            self._log_debug(f"Got token (first 20 chars): {self.token[:20]}...")
            return True
            
        except requests.exceptions.Timeout as e:
            error_message = (
                "Zeitueberschreitung bei der Authentifizierung: "
                f"{e}"
            )
            self._set_last_error("timeout", error_message)
            print(error_message)
            return False
        except requests.exceptions.RequestException as e:
            error_message = f"Netzwerkfehler bei der Authentifizierung: {e}"
            self._set_last_error("network", error_message)
            print(error_message)
            return False
        except (KeyError, TypeError) as e:
            error_message = f"Unerwartetes Antwortformat: {e}"
            self._set_last_error("response", error_message)
            print(error_message)
            self._log_debug(f"Response: {response.text}")
            return False

    def _graphql_request(self, query: str, variables: dict) -> dict:
        """Make an authenticated GraphQL request."""
        if not self.token:
            raise RuntimeError("Not authenticated. Call authenticate() first.")
        
        headers = {"Authorization": f"JWT {self.token}"}
        payload = {"query": query, "variables": variables}
        
        if self.debug:
            print("\n" + "="*80)
            print("GRAPHQL REQUEST:")
            print("="*80)
            print(f"URL: {GRAPHQL_URL}")
            print(f"Headers: {headers}")
            print(f"Payload: {json.dumps(payload, indent=2)}")
            print("="*80)
        
        try:
            response = self._post_with_retry(
                json_payload=payload,
                headers=headers,
            )
            
            if self.debug:
                print("\n" + "="*80)
                print("GRAPHQL RESPONSE:")
                print("="*80)
                print(f"Status: {response.status_code}")
                try:
                    response_data = response.json()
                    print(f"Body: {json.dumps(response_data, indent=2)}")
                except:
                    print(f"Body: {response.text}")
                print("="*80 + "\n")
            
            response.raise_for_status()
            data = response.json()
            
            if "errors" in data and not self.debug:
                error_message = f"GraphQL errors: {data['errors']}"
                self._set_last_error("graphql", error_message)
                print(error_message)
                # Return partial data if available
                return data.get("data", {})
            
            self._clear_last_error()
            return data.get("data", {})
            
        except requests.exceptions.Timeout as e:
            error_message = f"Zeitueberschreitung: {e}"
            self._set_last_error("timeout", error_message)
            print(error_message)
            return {}
        except requests.exceptions.RequestException as e:
            error_message = f"Netzwerkfehler: {e}"
            self._set_last_error("network", error_message)
            print(error_message)
            return {}

    def get_account_details(self, account_number: str) -> dict:
        """Get account details including meter information."""
        variables = {"accountNumber": account_number}
        result = self._graphql_request(ACCOUNT_DETAILS_QUERY, variables)
        return result.get("account", {})

    def get_accounts_from_viewer(self) -> list[dict]:
        """Get all accounts for the authenticated user using viewer query."""
        result = self._graphql_request(VIEWER_QUERY, {})
        viewer_data = result.get("viewer", {})
        accounts = viewer_data.get("accounts", [])
        self._log_debug(f"Found {len(accounts)} account(s) from viewer")
        return accounts

    def find_smart_meter(self, account_number: str) -> tuple[str, str, str] | None:
        """
        Find the smart meter ID and its parent property for an account.
        Returns tuple of (malo_number, meter_id, property_id) or None if not found.
        """
        variables = {"accountNumber": account_number}
        result = self._graphql_request(ACCOUNT_DETAILS_QUERY, variables)
        account_data = result.get("account", {})
        
        if not account_data:
            print("Keine Kontodaten gefunden")
            return None
        
        all_properties = account_data.get("allProperties", [])
        
        if not all_properties:
            print("Keine Eigenschaften für dieses Konto gefunden")
            return None
        
        # Look for smart meters in all properties
        for prop in all_properties:
            property_id = prop.get("id")
            malos = prop.get("electricityMalos", [])
            
            for malo in malos:
                malo_number = malo.get("maloNumber", "unknown")
                meter_data = malo.get("meter", {})
                
                if meter_data:
                    meter_id = meter_data.get("id")
                    should_receive_smart = meter_data.get("shouldReceiveSmartMeterData", False)
                    
                    if meter_id:
                        print(f"Zähler {meter_id} für MALO {malo_number} gefunden")
                        print(f"  - Eigenschafts-ID: {property_id}")
                        print(f"  - Soll Smart-Meter-Daten empfangen: {should_receive_smart}")
                        return (malo_number, meter_id, property_id)
        
        return None

    def get_active_tariff_agreement(self, account_number: str) -> TariffAgreement | None:
        result = self._graphql_request(ACCOUNT_DETAILS_QUERY, {"accountNumber": account_number})
        account_data = result.get("account", {})

        for prop in account_data.get("allProperties", []):
            for malo in prop.get("electricityMalos", []):
                for agreement in malo.get("agreements", []) or []:
                    if not agreement.get("isActive"):
                        continue
                    product = agreement.get("product") or {}
                    display_name = product.get("displayName")
                    if not display_name:
                        continue
                    return TariffAgreement(
                        display_name=display_name,
                        valid_from=agreement.get("validFrom", ""),
                        valid_to=agreement.get("validTo"),
                        agreement_id=agreement.get("id"),
                    )

        return None

    def get_tariff_settings_for_agreement(
        self,
        agreement: TariffAgreement,
    ) -> TariffSettings | None:
        if not agreement.agreement_id:
            return None

        result = self._graphql_request(
            RATE_STRUCTURE_QUERY,
            {"agreementId": agreement.agreement_id},
        )
        agreement_data = result.get("agreement", {})
        if not agreement_data:
            return None

        monthly_base_price_eur = _extract_monthly_base_price(agreement_data)
        if monthly_base_price_eur is None:
            print("[WARN] Konnte Grundpreis nicht aus API extrahieren, verwende Standardwert.")
            defaults = get_default_tariff_settings_for_type(TARIFF_INTELLIGENT_GO)
            monthly_base_price_eur = defaults.monthly_base_price_eur
        else:
            print(f"[INFO] Grundpreis aus API: {monthly_base_price_eur} EUR/Monat")

        return map_rate_structure_to_tariff_settings(
            agreement.display_name,
            agreement_data.get("unitRateInformation"),
            monthly_base_price_eur,
        )

    def get_tariff_rates_for_agreement(self, agreement: TariffAgreement) -> list[TariffRate]:
        if not agreement.agreement_id:
            return []

        result = self._graphql_request(
            RATE_STRUCTURE_QUERY,
            {"agreementId": agreement.agreement_id},
        )
        agreement_data = result.get("agreement", {})
        if not agreement_data:
            return []

        return _extract_tariff_rates(agreement_data.get("unitRateInformation"))

    def get_smart_usage(
        self,
        property_id: str,
        market_supply_point_id: str,
        day: datetime | date,
        reading_direction: str = "CONSUMPTION",
    ) -> list[dict]:
        """Fetch a single local day of hourly consumption via GetSmartUsage."""
        if isinstance(day, datetime):
            local_day = normalize_datetime(day).date()
        else:
            local_day = day

        start_local = datetime(local_day.year, local_day.month, local_day.day, tzinfo=APP_TIMEZONE)
        end_local = start_local + timedelta(days=1)
        start_utc = start_local.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        end_utc = end_local.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        variables = {
            "propertyId": property_id,
            "timezone": "Europe/Berlin",
            "startAt": start_utc,
            "endAt": end_utc,
            "utilityFilters": [
                {
                    "electricityFilters": {
                        "readingFrequencyType": "HOUR_INTERVAL",
                        "marketSupplyPointId": str(market_supply_point_id),
                        "readingDirection": _normalize_reading_direction(reading_direction),
                    }
                }
            ],
        }
        result = self._graphql_request(SMART_USAGE_QUERY, variables)
        property_data = result.get("property", {})
        measurements_data = property_data.get("measurements", {})
        edges = measurements_data.get("edges", [])

        intervals: list[dict] = []
        for edge in edges:
            node = edge.get("node", {})
            value = node.get("value")
            start_at = node.get("startAt")
            end_at = node.get("endAt")
            if value is None or not start_at or not end_at:
                continue

            try:
                start_time = normalize_datetime(datetime.fromisoformat(start_at))
                end_time = normalize_datetime(datetime.fromisoformat(end_at))
                direction = _extract_reading_direction(node)
                intervals.append(
                    {
                        "start": start_time,
                        "end": end_time,
                        "direction": direction,
                        "energy_kwh": float(value),
                        "consumption_kwh": float(value),
                        "net_kwh": float(value) if direction == "CONSUMPTION" else -float(value),
                        "duration_seconds": int((end_time - start_time).total_seconds()),
                        "unit": node.get("unit", "kWh"),
                        "source": node.get("source", "GetSmartUsage"),
                        "api_start": str(start_at),
                        "api_end": str(end_at),
                        "api_value": str(value),
                    }
                )
            except (ValueError, TypeError):
                continue

        intervals.sort(key=lambda item: item["start"])
        return intervals

    def get_consumption_smart_usage(
        self,
        property_id: str,
        market_supply_point_id: str,
        period_from: datetime | None = None,
        period_to: datetime | None = None,
        progress_callback=None,
        reading_direction: str = "CONSUMPTION",
    ) -> list[dict]:
        """Fetch hourly consumption day-by-day via GetSmartUsage."""
        today = get_today_start()
        yesterday_end = today - timedelta(seconds=1)

        if period_to and normalize_datetime(period_to) >= today:
            period_to = yesterday_end
            self._log_debug(f"Clamped period_to to yesterday: {period_to}")

        local_start_date = to_local_datetime(period_from).date() if period_from else to_local_datetime(yesterday_end).date()
        local_end_date = to_local_datetime(period_to).date() if period_to else local_start_date

        if local_end_date < local_start_date:
            return []

        all_intervals: list[dict] = []
        total_days = (local_end_date - local_start_date).days + 1
        for day_index in range(total_days):
            current_day = local_start_date + timedelta(days=day_index)
            day_intervals = self.get_smart_usage(
                property_id=property_id,
                market_supply_point_id=market_supply_point_id,
                day=current_day,
                reading_direction=reading_direction,
            )
            for interval in day_intervals:
                interval_start = normalize_datetime(interval["start"])
                interval_end = normalize_datetime(interval["end"])
                if period_from and interval_end <= normalize_datetime(period_from):
                    continue
                if period_to and interval_start > normalize_datetime(period_to):
                    continue
                all_intervals.append(interval)

            if progress_callback:
                progress_callback(len(all_intervals), day_index + 1, total_days, current_day)

        all_intervals.sort(key=lambda item: item["start"])
        return all_intervals

    def get_meter_reference_readings(
        self,
        account_number: str,
        meter_id: str | int,
    ) -> list[dict]:
        """Return reference meter readings from the API for calibration/UI selection."""
        cursor = None
        reference_readings: list[dict] = []

        while True:
            result = self._graphql_request(
                METER_READINGS_QUERY,
                {
                    "accountNumber": account_number,
                    "meterId": str(meter_id),
                    "cursor": cursor,
                },
            )
            readings_data = result.get("electricityMeterReadings", {})
            for edge in readings_data.get("edges", []):
                node = edge.get("node", {})
                read_at = node.get("readAt")
                value = node.get("value")
                type_of_read = node.get("typeOfRead")
                if not read_at or value is None or type_of_read not in {"INTERIM", "ESTIMATE"}:
                    continue

                try:
                    reference_readings.append(
                        {
                            "read_at": normalize_datetime(datetime.fromisoformat(read_at)),
                            "value": float(value),
                            "type_of_read": type_of_read,
                            "origin": node.get("origin"),
                            "register_obis_code": node.get("registerObisCode"),
                        }
                    )
                except (TypeError, ValueError):
                    continue

            page_info = readings_data.get("pageInfo", {})
            if not page_info.get("hasNextPage"):
                reference_readings.sort(key=lambda item: normalize_datetime(item["read_at"]), reverse=True)
                return reference_readings
            cursor = page_info.get("endCursor")
            if not cursor:
                reference_readings.sort(key=lambda item: normalize_datetime(item["read_at"]), reverse=True)
                return reference_readings

    def get_latest_interim_meter_reading(
        self,
        account_number: str,
        meter_id: str | int,
    ) -> dict | None:
        """Return the newest actual INTERIM meter reading for cumulative calibration."""
        for reading in self.get_meter_reference_readings(account_number, meter_id):
            if reading.get("type_of_read") == "INTERIM":
                return reading
        return None

    def get_consumption_graphql(self, property_id, period_from=None, period_to=None, fetch_all=False, progress_callback=None):
        """
        Get consumption data using GraphQL measurements query with hourly interval filter.
        
        Args:
            property_id: The property ID (from find_smart_meter)
            period_from: Start datetime (optional)
            period_to: End datetime (optional)
            fetch_all: If True, fetch all pages of data
            progress_callback: Optional callback function(current_count, page_num) to report progress
            
        Returns:
            List of consumption readings with start, end, and consumption_kwh
        """
        # Safety: Never fetch data for current day or future - data may be incomplete
        today = get_today_start()
        yesterday_end = today - timedelta(seconds=1)
        
        if period_to and normalize_datetime(period_to) >= today:
            period_to = yesterday_end
            self._log_debug(f"Clamped period_to to yesterday: {period_to}")
        
        all_intervals = []
        after_cursor = None
        page_count = 0
        max_pages = 100 if fetch_all else 1
        
        self._log_debug(f"Fetching measurements for property {property_id}")
        self._log_debug(f"fetch_all={fetch_all}, max_pages={max_pages}")
        
        while page_count < max_pages:
            page_count += 1
            
            # Build variables for the measurements query
            variables = {
                "propertyId": property_id,
                "first": 100,
                "utilityFilters": [{
                    "electricityFilters": {
                        "readingFrequencyType": "HOUR_INTERVAL"
                    }
                }],
                "timezone": "Europe/Berlin"
            }
            
            if after_cursor:
                variables["after"] = after_cursor
            
            # Add date filters if specified
            if period_from:
                variables["startAt"] = ensure_app_timezone(period_from).astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
            
            if period_to:
                variables["endAt"] = ensure_app_timezone(period_to).astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
            
            self._log_debug(f"Fetching page {page_count}, after={after_cursor}")
            
            result = self._graphql_request(MEASUREMENTS_QUERY, variables)
            
            if not result:
                self._log_debug("No response data, stopping")
                break
            
            property_data = result.get("property")
            if not property_data:
                self._log_debug(f"No property data in response: {result}")
                break
            
            measurements_data = property_data.get("measurements")
            if not measurements_data:
                self._log_debug(f"No measurements data: {property_data}")
                break
            
            edges = measurements_data.get("edges", [])
            page_info = measurements_data.get("pageInfo", {})
            
            self._log_debug(f"Page {page_count}: Got {len(edges)} edges")
            self._log_debug(f"Page info: hasNextPage={page_info.get('hasNextPage')}, endCursor={page_info.get('endCursor')}")
            
            if not edges:
                self._log_debug("No more edges, stopping pagination")
                break
            
            # Parse measurements from edges
            for edge in edges:
                node = edge.get("node", {})
                value = node.get("value")
                start_at = node.get("startAt")
                end_at = node.get("endAt")
                duration = node.get("durationInSeconds")
                direction = _extract_reading_direction(node)
                
                if value is not None and start_at and end_at:
                    try:
                        start_time = normalize_datetime(
                            datetime.fromisoformat(start_at.replace("Z", "+00:00"))
                        )
                        end_time = normalize_datetime(
                            datetime.fromisoformat(end_at.replace("Z", "+00:00"))
                        )
                        
                        all_intervals.append({
                            "start": start_time,
                            "end": end_time,
                            "direction": direction,
                            "energy_kwh": float(value),
                            "consumption_kwh": float(value),
                            "net_kwh": float(value) if direction == "CONSUMPTION" else -float(value),
                            "duration_seconds": duration,
                            "unit": node.get("unit", "kWh"),
                            "api_start": str(start_at),
                            "api_end": str(end_at),
                            "api_value": str(value),
                        })
                    except (ValueError, TypeError) as e:
                        self._log_debug(f"Error parsing measurement: {e}")
                        continue
            
            # Report progress
            if progress_callback:
                progress_callback(len(all_intervals), page_count)
            
            # Check if there are more pages
            if not page_info.get("hasNextPage"):
                self._log_debug("No more pages available")
                break
            
            after_cursor = page_info.get("endCursor")
            
            if not fetch_all:
                self._log_debug("fetch_all=False, stopping after first page")
                break
        
        # Sort by start time
        all_intervals.sort(key=lambda x: x["start"])
        
        self._log_debug(f"Total: Generated {len(all_intervals)} consumption intervals from {page_count} pages")
        
        return all_intervals


def to_local_datetime(dt: datetime) -> datetime:
    """Convert a datetime to a naive local app-timezone value for display."""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    if APP_TIMEZONE_HAS_IANA_DATA:
        return dt.astimezone(APP_TIMEZONE).replace(tzinfo=None)

    # Without IANA data, use the system local timezone conversion directly so
    # each timestamp gets the correct seasonal offset instead of a fixed one.
    return dt.astimezone().replace(tzinfo=None)


def format_datetime(dt: datetime, *, use_local_time: bool = False) -> str:
    """Format datetime for CSV output (European format: DD.MM.YYYY HH:MM:SS)."""
    dt = to_local_datetime(dt) if use_local_time else normalize_datetime(dt)
    return dt.strftime("%d.%m.%Y %H:%M:%S")


def ensure_app_timezone(dt: datetime) -> datetime:
    """Interpret naive datetimes in the app timezone and convert aware datetimes to it."""
    if dt.tzinfo is None:
        return dt.replace(tzinfo=APP_TIMEZONE)
    return dt.astimezone(APP_TIMEZONE)


def get_today_start() -> datetime:
    """Get the start of the current UTC day as a naive datetime."""
    return datetime.now(timezone.utc).replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
        tzinfo=None
    )


def normalize_datetime(dt: datetime) -> datetime:
    """Normalize datetimes to naive UTC values for storage and comparisons."""
    if dt.tzinfo is None:
        return dt
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def parse_date(date_str: str) -> datetime:
    """Parse date string in European format (DD.MM.YYYY)."""
    return datetime.strptime(date_str, "%d.%m.%Y")


def parse_datetime(datetime_str: str) -> datetime:
    """Parse datetime string in European format (DD.MM.YYYY HH:MM:SS)."""
    return datetime.strptime(datetime_str, "%d.%m.%Y %H:%M:%S")


def _reading_sort_value(raw_value) -> datetime:
    if isinstance(raw_value, datetime):
        return normalize_datetime(raw_value)
    return datetime.fromisoformat(str(raw_value))


def _resolve_meter_reading_offset(
    sorted_readings: list[dict],
    reference_reading: dict | None = None,
) -> float:
    """Return the cumulative offset before the first interval."""
    if not sorted_readings:
        return 0.0

    if reference_reading is None:
        for reading in sorted_readings:
            existing_value = reading.get("meter_reading_kwh")
            if existing_value is None:
                continue
            try:
                return float(existing_value) - float(reading.get("net_kwh", reading["consumption_kwh"]))
            except (TypeError, ValueError, KeyError):
                continue
        return 0.0

    if isinstance(reference_reading, list):
        reference_reading = next(
            (item for item in reference_reading if isinstance(item, dict)),
            None,
        )
        if reference_reading is None:
            return 0.0

    reference_time = reference_reading.get("read_at")
    reference_value = reference_reading.get("value")
    if not isinstance(reference_time, datetime):
        return 0.0

    try:
        anchor_value = float(reference_value)
    except (TypeError, ValueError):
        return 0.0

    normalized_reference_time = normalize_datetime(reference_time)
    consumption_until_anchor = 0.0
    for reading in sorted_readings:
        reading_end = _reading_sort_value(reading["end"])
        if reading_end <= normalized_reference_time:
            consumption_until_anchor += float(reading.get("net_kwh", reading["consumption_kwh"]))
        else:
            break

    return anchor_value - consumption_until_anchor


def build_readings_with_meter_reading(
    readings: list,
    reference_reading: dict | None = None,
) -> list[dict]:
    """Return sorted readings enriched with cumulative meter_reading_kwh."""
    if isinstance(reference_reading, list):
        reference_reading = next(
            (item for item in reference_reading if isinstance(item, dict)),
            None,
        )
    sorted_readings = sorted(
        readings,
        key=lambda reading: _reading_sort_value(reading["start"]),
    )

    running_meter_reading = _resolve_meter_reading_offset(
        sorted_readings,
        reference_reading=reference_reading,
    )
    enriched_readings: list[dict] = []
    for reading in sorted_readings:
        net_value = float(reading.get("net_kwh", reading["consumption_kwh"]))
        running_meter_reading += net_value
        enriched_reading = dict(reading)
        enriched_reading["meter_reading_kwh"] = round(running_meter_reading, 4)
        enriched_reading.setdefault("direction", "CONSUMPTION")
        enriched_reading.setdefault("energy_kwh", abs(float(reading["consumption_kwh"])))
        enriched_reading.setdefault("net_kwh", net_value)
        enriched_readings.append(enriched_reading)

    return enriched_readings


def convert_readings_for_export(
    readings: list,
    reference_reading: dict | None = None,
    *,
    use_local_time: bool = False,
) -> list:
    """Convert readings to serializable format for JSON/YAML export."""
    export_data = []
    for reading in build_readings_with_meter_reading(readings, reference_reading=reference_reading):
        export_data.append({
            'start': (
                to_local_datetime(reading['start']).isoformat()
                if isinstance(reading['start'], datetime) and use_local_time
                else reading['start'].isoformat() if isinstance(reading['start'], datetime) else reading['start']
            ),
            'end': (
                to_local_datetime(reading['end']).isoformat()
                if isinstance(reading['end'], datetime) and use_local_time
                else reading['end'].isoformat() if isinstance(reading['end'], datetime) else reading['end']
            ),
            'direction': reading.get('direction', 'CONSUMPTION'),
            'energy_kwh': abs(float(reading.get('energy_kwh', reading['consumption_kwh']))),
            'consumption_kwh': reading['consumption_kwh'],
            'net_kwh': reading.get('net_kwh', float(reading['consumption_kwh'])),
            'meter_reading_kwh': reading['meter_reading_kwh'],
            'duration_seconds': reading.get('duration_seconds'),
            'unit': reading.get('unit', 'kWh'),
            'api_start': reading.get('api_start'),
            'api_end': reading.get('api_end'),
            'api_value': reading.get('api_value'),
        })
    return export_data


def save_to_json(
    readings: list,
    output_path: Path,
    reference_reading: dict | None = None,
    *,
    use_local_time: bool = False,
) -> bool:
    """Save readings to JSON format."""
    try:
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'total_readings': len(readings),
                'source': 'Octopus Energy Germany Smart Meter'
            },
            'readings': convert_readings_for_export(
                readings,
                reference_reading=reference_reading,
                use_local_time=use_local_time,
            )
        }
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Fehler beim Speichern als JSON: {e}")
        return False


def save_to_yaml(
    readings: list,
    output_path: Path,
    reference_reading: dict | None = None,
    *,
    use_local_time: bool = False,
) -> bool:
    """Save readings to YAML format."""
    try:
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'total_readings': len(readings),
                'source': 'Octopus Energy Germany Smart Meter'
            },
            'readings': convert_readings_for_export(
                readings,
                reference_reading=reference_reading,
                use_local_time=use_local_time,
            )
        }
        with open(output_path, 'w', encoding='utf-8') as f:
            yaml.dump(export_data, f, allow_unicode=True, sort_keys=False)
        return True
    except Exception as e:
        print(f"Fehler beim Speichern als YAML: {e}")
        return False


def fill_excel_template(
    readings: list,
    template_path: str,
    output_path: str,
    tariff_go_ct: float = DEFAULT_TARIFF_GO_CT,
    tariff_standard_ct: float = DEFAULT_TARIFF_STANDARD_CT,
    tariff_high_ct: float = 0.0,
    monthly_base_price_eur: float = DEFAULT_MONTHLY_BASE_PRICE_EUR,
    tariff_type: str | None = None,
):
    """
    Fill a German electricity tariff Excel template with consumption data.
    
    Updates:
    1. Einstellungen sheet: B5 = first date, B6 = last date from CSV
    2. Verbrauch sheet: Only column C (rows 9+) where date/hour match and cell is empty
    
    Args:
        readings: List of consumption readings
        template_path: Path to the Excel template
        output_path: Path for the output file
    """
    try:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        print("Fehler: openpyxl ist für Excel-Unterstützung erforderlich")
        print("Installieren mit: pip install openpyxl")
        return False
    
    try:
        requested_tariff_type = tariff_type or TARIFF_INTELLIGENT_GO
        template_path_obj = Path(template_path)
        output_path_obj = Path(output_path)
        workbook_tariff_type = requested_tariff_type
        bundled_template = get_bundled_excel_template_path(requested_tariff_type)
        try:
            output_path_obj.parent.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
        if template_path_obj.exists():
            workbook_tariff_type = detect_excel_template_type(template_path_obj)
            if tariff_type and workbook_tariff_type != tariff_type:
                print(
                    "Fehler: Die ausgewaehlte Excel-Datei verwendet ein anderes Tarifzonenmodell "
                    f"({workbook_tariff_type}) als der aktuelle Tarif ({tariff_type})."
                )
                return False
        elif bundled_template.exists():
            try:
                template_path_obj.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(bundled_template, template_path_obj)
                print(f"Excel-Vorlage aus dem Paket kopiert nach: {template_path_obj}")
            except OSError:
                pass
        elif requested_tariff_type == TARIFF_INTELLIGENT_HEAT:
            stock_template = get_bundled_excel_template_path(TARIFF_INTELLIGENT_GO)
            if stock_template.exists():
                create_heat_excel_template(stock_template, template_path_obj)
                print(f"Heat-Excel-Vorlage erzeugt nach: {template_path_obj}")
        elif not template_path_obj.exists():
            print(f"Fehler: Excel-Vorlage nicht gefunden: {template_path_obj}")
            return False

        layout = get_excel_layout(workbook_tariff_type)
        template_path = str(template_path_obj)
        output_path = str(output_path_obj)

        # Create backup of original file
        backup_path = template_path + ".backup"
        shutil.copy2(template_path, backup_path)
        print(f"Sicherung erstellt: {backup_path}")
        
        # Only preserve VBA for macro-enabled workbooks. Using keep_vba=True on
        # a normal .xlsx can make Excel reject the saved file even though
        # LibreOffice still opens it.
        keep_vba = template_path_obj.suffix.lower() in {".xlsm", ".xltm"}
        wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=keep_vba)

        if keep_vba:
            print(f"Vorlage geladen: {template_path} (mit VBA/Makros erhalten)")
        else:
            print(f"Vorlage geladen: {template_path}")
        
        # Get the sheets
        ws_verbrauch = wb['Verbrauch']  # Consumption sheet
        ws_einstellungen = wb['Einstellungen']  # Settings sheet
        
        # Create a dictionary of readings by datetime for quick lookup
        # Key: (date, hour) tuple, Value: consumption_kwh
        readings_by_datetime = {}
        for reading in readings:
            start = reading["start"]
            date_key = start.strftime("%Y-%m-%d")
            hour_key = start.hour  # 0-23
            readings_by_datetime[(date_key, hour_key)] = reading["consumption_kwh"]
        
        print(f"{len(readings_by_datetime)} stündliche Einträge zum Abgleich vorhanden")
        
        # Get date range from CSV
        csv_dates = sorted(set(k[0] for k in readings_by_datetime.keys()))
        first_date = csv_dates[0]
        last_date = csv_dates[-1]
        print(f"CSV-Datumsbereich: {first_date} bis {last_date}")
        
        # Update Einstellungen sheet based on the detected template layout.
        print(f"\nUpdating Einstellungen sheet:")
        ws_einstellungen[layout["tariff_low"]].value = float(tariff_go_ct)
        ws_einstellungen[layout["tariff_standard"]].value = float(tariff_standard_ct)
        if workbook_tariff_type == TARIFF_INTELLIGENT_HEAT and layout["tariff_high"]:
            ws_einstellungen[layout["tariff_high"]].value = float(tariff_high_ct)
        ws_einstellungen[layout["base_price"]].value = float(monthly_base_price_eur)
        first_date_dt = datetime.strptime(first_date, "%Y-%m-%d")
        last_date_dt = datetime.strptime(last_date, "%Y-%m-%d")

        # Format as German date (DD.MM.YYYY) for the Excel
        ws_einstellungen[layout["start_date"]].value = first_date_dt
        ws_einstellungen[layout["end_date"]].value = last_date_dt
        print(f"  {layout['tariff_low']} (Tarif Niedrig/Go): {tariff_go_ct}")
        print(f"  {layout['tariff_standard']} (Tarif Standard): {tariff_standard_ct}")
        if workbook_tariff_type == TARIFF_INTELLIGENT_HEAT and layout["tariff_high"]:
            print(f"  {layout['tariff_high']} (Tarif Hoch): {tariff_high_ct}")
        print(f"  {layout['start_date']} (Start): {first_date}")
        print(f"  {layout['end_date']} (Ende): {last_date}")
        print(f"  {layout['base_price']} (Grundpreis): {monthly_base_price_eur}")
        
        # Template structure for Verbrauch sheet
        DATA_START_ROW = 9
        DATE_COL = 1  # Column A
        HOUR_COL = 2  # Column B
        CONSUMPTION_COL = 3  # Column C
        
        # Fill consumption values in Verbrauch sheet
        # The formulas in A and B will auto-calculate based on the template start date cell.
        # We only need to fill column C where we have matching data and cell is empty
        print(f"\nFilling Verbrauch sheet (column C from row {DATA_START_ROW})...")
        
        filled_count = 0
        rows_checked = 0
        
        # Calculate how many rows we need to check (24 hours * number of days in range)
        days_count = (last_date_dt - first_date_dt).days + 1
        expected_rows = days_count * 24
        max_row = DATA_START_ROW + expected_rows + 48  # Add buffer
        
        for row in range(DATA_START_ROW, min(max_row + 1, ws_verbrauch.max_row + 1)):
            rows_checked += 1
            
            # Get date from column A (calculated from the template start date cell)
            date_cell = ws_verbrauch.cell(row=row, column=DATE_COL)
            if isinstance(date_cell, MergedCell):
                continue
            
            # Mirror the workbook formula logic in Python while filling values.
            days_offset = (row - DATA_START_ROW) // 24
            date_parsed = first_date_dt + timedelta(days=days_offset)
            
            # Get hour from column B
            # Formula: =MOD(ROW()-9,24)
            hour_parsed = (row - DATA_START_ROW) % 24
            
            # Create lookup key
            date_key = date_parsed.strftime("%Y-%m-%d")
            lookup_key = (date_key, hour_parsed)
            
            # Check if we have data for this date/hour
            if lookup_key in readings_by_datetime:
                consumption_cell = ws_verbrauch.cell(row=row, column=CONSUMPTION_COL)
                if isinstance(consumption_cell, MergedCell):
                    continue
                
                # Only write if cell is empty
                current_value = consumption_cell.value
                if current_value is None or current_value == '' or current_value == 0:
                    consumption_cell.value = readings_by_datetime[lookup_key]
                    filled_count += 1
            
            # Progress update
            if rows_checked % 1000 == 0:
                print(f"  {rows_checked} Zeilen geprüft, {filled_count} Werte gefüllt...")
        
        print(f"{filled_count} Verbrauchswerte in {rows_checked} Zeilen gefüllt")
        
        # Save the workbook
        wb.save(output_path)
        print(f"Excel-Datei gespeichert nach: {output_path}")
        return True
        
    except Exception as e:
        print(f"Fehler beim Füllen der Excel-Vorlage: {e}")
        import traceback
        traceback.print_exc()
        return False


def read_existing_csv(csv_path: Path) -> tuple[list, datetime | None]:
    """
    Read existing consumption.csv and return data plus latest interval end.
    
    Returns:
        Tuple of (existing_data, latest_interval_end)
    """
    if not csv_path.exists():
        return [], None
    
    existing_data = []
    latest_interval_end = None
    
    try:
        with open(csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    # Try European format first (DD.MM.YYYY HH:MM:SS), then ISO
                    try:
                        start = datetime.strptime(row['start'], "%d.%m.%Y %H:%M:%S")
                        end = datetime.strptime(row['end'], "%d.%m.%Y %H:%M:%S")
                    except ValueError:
                        # Fallback to ISO format for backwards compatibility
                        start = normalize_datetime(datetime.fromisoformat(row['start']))
                        end = normalize_datetime(datetime.fromisoformat(row['end']))
                    direction = _normalize_reading_direction(row.get('direction'))
                    energy = row.get('energy_kwh', row.get('consumption_kwh'))
                    net_value = row.get('net_kwh')
                    if net_value in (None, ''):
                        net_value = float(energy) if direction == "CONSUMPTION" else -float(energy)
                    consumption = float(row.get('consumption_kwh', energy))
                    
                    reading = {
                        'start': start,
                        'end': end,
                        'direction': direction,
                        'energy_kwh': abs(float(energy)),
                        'consumption_kwh': consumption
                    }
                    reading['net_kwh'] = float(net_value)
                    if "T" in row['start']:
                        reading['api_start'] = row['start']
                    if "T" in row['end']:
                        reading['api_end'] = row['end']
                    if row.get('consumption_kwh') not in (None, ''):
                        reading['api_value'] = row['consumption_kwh']
                    meter_reading = row.get('meter_reading_kwh')
                    if meter_reading not in (None, ''):
                        try:
                            reading['meter_reading_kwh'] = float(meter_reading)
                        except ValueError:
                            pass

                    existing_data.append(reading)
                    
                    if latest_interval_end is None or end > latest_interval_end:
                        latest_interval_end = end
                        
                except (KeyError, ValueError) as e:
                    continue
        
        print(f"Read {len(existing_data)} existing readings from {csv_path}")
        if latest_interval_end:
            print(f"Latest interval end in CSV: {latest_interval_end}")
        
        return existing_data, latest_interval_end
        
    except Exception as e:
        print(f"Error reading existing CSV: {e}")
        return [], None


def read_existing_json(json_path: Path) -> tuple[list, datetime | None]:
    """Read cached readings.json and return data plus latest interval end."""
    if not json_path.exists():
        return [], None

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    except Exception as e:
        print(f"Error reading existing JSON: {e}")
        return [], None

    raw_readings = payload.get('readings', []) if isinstance(payload, dict) else []
    existing_data = []
    latest_interval_end = None

    for row in raw_readings:
        try:
            start = normalize_datetime(datetime.fromisoformat(row['start']))
            end = normalize_datetime(datetime.fromisoformat(row['end']))
            direction = _normalize_reading_direction(row.get('direction'))
            energy = row.get('energy_kwh', row.get('consumption_kwh'))
            net_value = row.get('net_kwh')
            if net_value in (None, ''):
                net_value = float(energy) if direction == "CONSUMPTION" else -float(energy)
            reading = {
                'start': start,
                'end': end,
                'direction': direction,
                'energy_kwh': abs(float(energy)),
                'consumption_kwh': float(row['consumption_kwh']),
            }
            reading['net_kwh'] = float(net_value)
            if row.get('api_start'):
                reading['api_start'] = str(row['api_start'])
            if row.get('api_end'):
                reading['api_end'] = str(row['api_end'])
            if row.get('api_value') not in (None, ''):
                reading['api_value'] = str(row['api_value'])
            meter_reading = row.get('meter_reading_kwh')
            if meter_reading not in (None, ''):
                reading['meter_reading_kwh'] = float(meter_reading)
            existing_data.append(reading)
            if latest_interval_end is None or end > latest_interval_end:
                latest_interval_end = end
        except (KeyError, TypeError, ValueError):
            continue

    print(f"Read {len(existing_data)} existing readings from {json_path}")
    if latest_interval_end:
        print(f"Latest interval end in JSON: {latest_interval_end}")

    return existing_data, latest_interval_end


def write_readings_json(readings: list[dict], json_path: Path) -> bool:
    """Persist calibrated readings to the internal JSON cache."""
    try:
        json_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass

    try:
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'total_readings': len(readings),
                'source': 'Octopus Energy Germany Smart Meter',
            },
            'readings': convert_readings_for_export(readings),
        }
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Fehler beim Speichern als readings.json: {e}")
        return False


def write_consumption_csv(readings: list[dict], csv_path: Path) -> bool:
    """Persist readings as API-style CSV with raw timestamps and value precision."""
    try:
        csv_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass

    try:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['start', 'end', 'direction', 'energy_kwh', 'consumption_kwh', 'net_kwh'])
            for reading in readings:
                writer.writerow([
                    reading.get('api_start')
                    or (reading['start'].replace(tzinfo=timezone.utc).isoformat() if isinstance(reading['start'], datetime) else reading['start']),
                    reading.get('api_end')
                    or (reading['end'].replace(tzinfo=timezone.utc).isoformat() if isinstance(reading['end'], datetime) else reading['end']),
                    reading.get('direction', 'CONSUMPTION'),
                    reading.get('energy_kwh', abs(float(reading.get('consumption_kwh', 0.0)))),
                    reading.get('api_value')
                    or str(reading['consumption_kwh']),
                    reading.get('net_kwh', reading.get('consumption_kwh', 0.0)),
                ])
        return True
    except Exception as e:
        print(f"Fehler beim Speichern als consumption.csv: {e}")
        return False


def consumption_csv_has_api_format(csv_path: Path) -> bool:
    """Return True when consumption.csv stores timestamps in API-style ISO format."""
    if not csv_path.exists():
        return False

    try:
        with open(csv_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            first_row = next(reader, None)
    except Exception:
        return False

    if not first_row:
        return False

    start_value = str(first_row.get('start', ''))
    end_value = str(first_row.get('end', ''))
    return 'T' in start_value and ('+' in start_value or start_value.endswith('Z')) and 'T' in end_value


def migrate_consumption_csv_to_config() -> Path:
    """Move any legacy Documents consumption.csv into the config folder."""
    config_csv = get_default_output_path()
    legacy_csv = get_smartmeter_data_folder() / "consumption.csv"

    try:
        os.makedirs(config_csv.parent, exist_ok=True)
    except Exception:
        pass

    if not legacy_csv.exists():
        return config_csv

    if config_csv.exists():
        try:
            legacy_csv.unlink()
        except OSError:
            pass
        return config_csv

    legacy_data, _ = read_existing_csv(legacy_csv)
    if legacy_data:
        write_readings_json(legacy_data, config_csv)

    try:
        legacy_csv.unlink()
    except OSError:
        pass

    return config_csv


def load_existing_consumption_data() -> tuple[list, datetime | None]:
    """
    Load existing consumption data.

    Returns:
        Tuple of (readings, latest_interval_end)
    """
    config_csv = migrate_consumption_csv_to_config()
    return read_existing_json(config_csv)


def merge_readings(all_data: list, reference_reading: dict | None = None) -> list[dict]:
    """Merge all data, remove duplicates, sort by time, and calibrate readings."""
    if not all_data:
        return []

    seen = {}
    for reading in all_data:
        key = reading['start'].isoformat()
        seen[key] = reading

    unique_data = list(seen.values())
    unique_data.sort(key=lambda x: x['start'])
    return build_readings_with_meter_reading(unique_data, reference_reading=reference_reading)


def readings_changed(existing_readings: list[dict], merged_readings: list[dict]) -> bool:
    """Return True when the cached readings are empty or differ from merged data."""
    if not existing_readings:
        return bool(merged_readings)
    if len(existing_readings) != len(merged_readings):
        return True

    for current, merged in zip(existing_readings, merged_readings):
        current_start = normalize_datetime(current['start'])
        merged_start = normalize_datetime(merged['start'])
        current_end = normalize_datetime(current['end'])
        merged_end = normalize_datetime(merged['end'])
        if current_start != merged_start:
            return True
        if current_end != merged_end:
            return True
        if float(current.get('net_kwh', current['consumption_kwh'])) != float(merged.get('net_kwh', merged['consumption_kwh'])):
            return True

    return False


def save_data(
    all_data: list,
    output_path: Path,
    output_format: str = "csv",
    reference_reading: dict | None = None,
    *,
    use_local_time: bool = False,
):
    """
    Save data to the specified format.
    
    Args:
        all_data: List of all readings
        output_path: Base path (without extension)
        output_format: One of 'csv', 'json', 'yaml'
    """
    if not all_data:
        print("Keine Daten zum Speichern")
        return False
    
    # Remove duplicates based on start time (keep last occurrence)
    seen = {}
    for reading in all_data:
        key = reading['start'].isoformat() if isinstance(reading['start'], datetime) else reading['start']
        seen[key] = reading
    
    # Convert back to list and sort
    unique_data = list(seen.values())
    unique_data.sort(key=lambda x: x['start'] if isinstance(x['start'], datetime) else datetime.fromisoformat(x['start']))
    export_rows = build_readings_with_meter_reading(unique_data, reference_reading=reference_reading)
    
    if output_format == "csv":
        # Change extension to .csv
        csv_path = output_path.with_suffix('.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['start', 'end', 'direction', 'energy_kwh', 'consumption_kwh', 'net_kwh', 'meter_reading_kwh'])
            for reading in export_rows:
                writer.writerow([
                    format_datetime(reading['start'], use_local_time=use_local_time) if isinstance(reading['start'], datetime) else reading['start'],
                    format_datetime(reading['end'], use_local_time=use_local_time) if isinstance(reading['end'], datetime) else reading['end'],
                    reading.get('direction', 'CONSUMPTION'),
                    reading.get('energy_kwh', abs(float(reading['consumption_kwh']))),
                    reading['consumption_kwh'],
                    reading.get('net_kwh', reading['consumption_kwh']),
                    reading['meter_reading_kwh'],
                ])
        print(f"{len(unique_data)} Einträge gespeichert nach {csv_path}")
        return True
    
    elif output_format == "json":
        json_path = output_path.with_suffix('.json')
        if save_to_json(
            unique_data,
            json_path,
            reference_reading=reference_reading,
            use_local_time=use_local_time,
        ):
            print(f"{len(unique_data)} Einträge als JSON gespeichert nach {json_path}")
            return True
        return False
    
    elif output_format == "yaml":
        yaml_path = output_path.with_suffix('.yaml')
        if save_to_yaml(
            unique_data,
            yaml_path,
            reference_reading=reference_reading,
            use_local_time=use_local_time,
        ):
            print(f"{len(unique_data)} Einträge als YAML gespeichert nach {yaml_path}")
            return True
        return False
    
    else:
        print(f"Unbekanntes Format: {output_format}")
        return False


def main():
    parser = argparse.ArgumentParser(
        description="Fetch smart meter data from Octopus Energy Germany"
    )
    parser.add_argument(
        "--email", 
        required=True,
        help="Your Octopus Energy email"
    )
    parser.add_argument(
        "--password", 
        required=True,
        help="Your Octopus Energy password"
    )
    parser.add_argument(
        "--account-number", 
        required=False,
        help="Your account number (e.g., A-12345678). If not provided, will be auto-discovered."
    )
    parser.add_argument(
        "--meter-id", 
        help="Smart meter ID (optional - will auto-discover if not provided)"
    )
    parser.add_argument(
        "--property-id", 
        help="Property ID (optional - will auto-discover if not provided)"
    )
    parser.add_argument(
        "--output", 
        default="",
        help="Optionaler Zusatzexportpfad fuer CSV/JSON/YAML"
    )
    parser.add_argument(
        "--period-from", 
        help="Start date (DD.MM.YYYY)"
    )
    parser.add_argument(
        "--period-to", 
        help="End date (DD.MM.YYYY)"
    )
    parser.add_argument(
        "--format", 
        choices=["csv", "hourly", "all"],
        default="csv",
        help="Data format: csv (raw intervals), hourly (interpolated), all (all columns)"
    )
    parser.add_argument(
        "--output-format",
        choices=["csv", "json", "yaml"],
        default="csv",
        help="Output file format (default: csv)"
    )
    parser.add_argument(
        "--fill-excel",
        metavar="TEMPLATE",
        default="",
        help="Excel-Datei/Vorlage mit Verbrauchsdaten fuellen"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug output for all API requests"
    )
    parser.add_argument(
        "--list-accounts",
        action="store_true",
        help="List all accounts for the authenticated user and exit"
    )
    
    args = parser.parse_args()
    
    # Handle --list-accounts option
    if args.list_accounts:
        if not args.email or not args.password:
            print("Fehler: E-Mail und Passwort sind erforderlich, um Konten aufzulisten.")
            print("Verwendung: --email user@example.com --password pass --list-accounts")
            sys.exit(1)
        
        client = OctopusGermanyClient(args.email, args.password, debug=args.debug)
        print("Authentifizierung...")
        if not client.authenticate():
            print("Authentifizierung fehlgeschlagen!")
            sys.exit(1)
        
        print("\nVerfügbare Konten:")
        accounts = client.get_accounts_from_viewer()
        if accounts:
            for acc in accounts:
                acc_num = acc.get('number', 'unknown')
                acc_id = acc.get('id', 'unknown')
                print(f"  - Kundennummer: {acc_num} (ID: {acc_id})")
        else:
            print("  Keine Konten gefunden.")
        sys.exit(0)
    
    # Initialize app config folder once at startup
    config_ok, config_error = init_app_config_folder()
    if not config_ok:
        print(f"Warnung: {config_error}")
        print("Einstellungen werden nicht gespeichert.")
    cleanup_app_config_folder()

    # Ensure the data directory exists before continuing.
    ensure_smartmeter_data_folder()
    
    # Parse dates
    period_from = None
    period_to = None
    
    if args.period_from:
        period_from = parse_date(args.period_from)
        print(f"Zeitraum von: {period_from}")
    
    if args.period_to:
        period_to = parse_date(args.period_to) + timedelta(days=1) - timedelta(seconds=1)
        print(f"Zeitraum bis: {period_to}")
    
    # Read existing data first (before any authentication)
    output_path = Path(args.output) if args.output else None
    if output_path is not None:
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
    existing_data, latest_interval_end = load_existing_consumption_data()
    consumption_csv_path = get_default_consumption_csv_path()
    csv_needs_api_refresh = existing_data and not consumption_csv_has_api_format(consumption_csv_path)
    
    # Determine date range for fetching new data
    # Never fetch data for the current day (data may be incomplete)
    today = get_today_start()
    yesterday = today - timedelta(days=1)
    yesterday_end = yesterday + timedelta(days=1) - timedelta(seconds=1)
    
    fetch_from = period_from
    fetch_to = period_to
    
    # Clamp fetch_to to yesterday if not specified or if it's today or in the future
    if fetch_to is None or fetch_to >= today:
        fetch_to = yesterday_end
        print(f"\nNote: Limiting data fetch to yesterday ({yesterday.date()}) - current day data excluded")
    
    if not fetch_from and latest_interval_end:
        # Start from the latest data we have (minus 1 hour overlap to be safe)
        fetch_from = latest_interval_end - timedelta(hours=1)
        # But don't go beyond yesterday
        if fetch_from > yesterday_end:
            fetch_from = yesterday_end
        print(f"\nAuto-detected: Fetching new data from {fetch_from}")
    
    # Check if we actually need to fetch data
    # If we have existing data and the user just wants to fill Excel, skip fetching
    need_to_fetch = True
    if existing_data and not args.period_from and not args.period_to:
        # We have data and no specific date range requested by user
        # Hourly intervals are [start, end), so yesterday is complete only if the
        # last interval ends at or after today's midnight.
        if latest_interval_end and latest_interval_end >= today:
            print(f"\nCSV already has complete data up to {yesterday.date()}")
            print("Keine neuen Daten abzurufen. Verwenden Sie --period-from für erzwungenen Abruf.")
            need_to_fetch = False
            # Reset fetch_from since we don't need to fetch
            fetch_from = None
    if csv_needs_api_refresh and existing_data:
        need_to_fetch = True
        fetch_from = min(reading["start"] for reading in existing_data)
        print("\nconsumption.csv verwendet noch nicht das API-Format. Lade Daten neu fuer den CSV-Cache.")
    
    # Initialize client and authenticate only if we need to fetch
    client = None
    property_id = args.property_id
    account_number = args.account_number
    meter_id = args.meter_id
    malo_number = None
    reference_reading = None
    
    if need_to_fetch:
        client = OctopusGermanyClient(args.email, args.password, debug=args.debug)
        
        # Authenticate
        print("\nAuthenticating...")
        if not client.authenticate():
            print("Authentifizierung fehlgeschlagen!")
            sys.exit(1)
        print("Authentication successful!")
        
        # Auto-discover account number if not provided
        if not account_number:
            print("\nDiscovering account number...")
            accounts = client.get_accounts_from_viewer()
            if not accounts:
                print("Kein Konto gefunden! Überprüfen Sie Ihre Zugangsdaten.")
                sys.exit(1)
            if len(accounts) > 1:
                print(f"Mehrere Konten gefunden ({len(accounts)}). Bitte wählen Sie eines mit --account-number:")
                for acc in accounts:
                    print(f"  - {acc.get('number', 'unknown')} (ID: {acc.get('id', 'unknown')})")
                sys.exit(1)
            account_number = accounts[0].get('number')
            print(f"Gefundene Kundennummer: {account_number}")
        
        # Get meter / MaLo details for GetSmartUsage
        if not property_id or not malo_number:
            print("\nDiscovering smart meter...")
            meter_info = client.find_smart_meter(account_number)
            if meter_info:
                discovered_malo_number, discovered_meter_id, discovered_property_id = meter_info
                malo_number = discovered_malo_number
                meter_id = discovered_meter_id
                if not property_id:
                    property_id = discovered_property_id
                print(f"Verwende Zähler-ID: {meter_id}")
                print(f"Verwende Eigenschafts-ID: {property_id}")
            else:
                print("Kein Smart Meter gefunden!")
                print("\nTroubleshooting:")
                print("1. Prüfen Sie, ob ein Smart Meter installiert ist")
                print("2. Vergewissern Sie sich, dass der Zähler für Smart-Readings freigeschaltet ist")
                print("3. Kontaktieren Sie Octopus Energy, falls das Problem weiterhin besteht")
                sys.exit(1)
    
    # Fetch consumption data only if needed
    new_readings = []
    if need_to_fetch:
        if account_number and meter_id:
            reference_reading = client.get_latest_interim_meter_reading(account_number, meter_id)
            if reference_reading:
                print(
                    "Verwende INTERIM-Zaehlerstand als Referenz: "
                    f"{reference_reading['value']:.3f} kWh "
                    f"am {format_datetime(reference_reading['read_at'])}"
                )

        print("\nFetching consumption data...")
        if fetch_from or fetch_to:
            print(f"Datumsbereich: {fetch_from or 'alle'} bis {fetch_to or 'alle'}")
        else:
            print("Alle verfügbaren Daten werden abgerufen...")
        
        # Progress callback for CLI
        def cli_progress(count, page):
            print(f"  Empfange... {count} Einträge (Seite {page})", end='\r', flush=True)
        
        new_readings = client.get_consumption_graphql(
            property_id,
            period_from=fetch_from,
            period_to=fetch_to,
            fetch_all=True,
            progress_callback=cli_progress
        )
        if new_readings:
            print()  # New line after progress
    
    if not new_readings and not existing_data:
        print("\nKeine Verbrauchsdaten gefunden!")
        print("\nMögliche Gründe:")
        print("- Der Stromzähler hat noch keine Smart-Meter Funktionalität")
        print("- Es wurden keine Werte gefunden")
        print("- Das Datum ist außerhalb der vorhandenen Daten")
        sys.exit(1)
    
    if new_readings:
        print(f"\n{len(new_readings)} neue Verbrauchsdatenintervalle gelesen")
        
        # Calculate total consumption
        total_kwh = sum(r["consumption_kwh"] for r in new_readings)
        print(f"Gesamtverbrauch neuer Daten: {total_kwh:.2f} kWh")
    else:
        print("\nEs wurden keine neuen Daten gefunden (alles bereits im Cache enthalten)")
    
    # Merge existing and new data
    all_readings = existing_data + new_readings
    
    if not all_readings:
        print("Keine Daten zum Speichern gefunden!")
        sys.exit(1)
    
    # Remove duplicates and save
    print(f"\nFasse Daten zusammen: {len(existing_data)} existierende + {len(new_readings)} neue = {len(all_readings)} total")

    config_cache_path = get_default_output_path()
    final_data = merge_readings(all_readings, reference_reading=reference_reading)
    if readings_changed(existing_data, final_data):
        if not write_readings_json(final_data, config_cache_path):
            print("Fehler beim Speichern von readings.json")
            sys.exit(1)
        if not write_consumption_csv(final_data, consumption_csv_path):
            print("Fehler beim Speichern von consumption.csv")
            sys.exit(1)
        print(f"{len(final_data)} Eintraege gespeichert nach {config_cache_path}")
        print(f"{len(final_data)} Eintraege gespeichert nach {consumption_csv_path}")
    elif csv_needs_api_refresh:
        if not write_consumption_csv(final_data, consumption_csv_path):
            print("Fehler beim Speichern von consumption.csv")
            sys.exit(1)
        print(f"{len(final_data)} Eintraege gespeichert nach {consumption_csv_path}")
    else:
        print(f"Keine neuen Daten fuer {config_cache_path.name}; Cache bleibt unveraendert.")

    # Optional extra export only when explicitly requested
    output_format = args.output_format
    if output_path is not None:
        save_data(final_data, output_path, output_format, reference_reading=reference_reading)
    
    # Show data summary
    print(f"\nInsgesamt: {len(final_data)} Verbrauchsdaten")
    if final_data:
        total_kwh = sum(r["consumption_kwh"] for r in final_data)
        print(f"Gesamtverbrauch im Cache: {total_kwh:.2f} kWh")
        
        # Show data granularity
        if len(final_data) > 1:
            durations = []
            for i in range(1, min(10, len(final_data))):
                duration = (final_data[i]["start"] - final_data[i-1]["start"]).total_seconds() / 3600
                durations.append(duration)
            avg_duration = sum(durations) / len(durations)
            print(f"Durchschnittliches Intervall: {avg_duration:.1f} Stunden")
    
    # Fill Excel template if requested
    if args.fill_excel:
        template_path = Path(args.fill_excel)
        if not template_path.exists():
            print(f"\nFehler: Excel template nicht gefunden: {template_path}")
        else:
            print(f"\nExcel wird mittels csv-Daten befüllt: {template_path}")
            success = fill_excel_template(
                final_data, 
                str(template_path), 
                str(template_path)  # Save in-place
            )
            if not success:
                sys.exit(1)
    
    print("\n" + "="*60)
    print("Daten erfolgreich geschrieben nach:")
    print(f"  - readings.json: {config_cache_path}")
    print(f"  - consumption.csv: {consumption_csv_path}")
    if output_path is not None and output_format == "csv":
        print(f"  - Zusatz-CSV: {output_path.with_suffix('.csv')}")
    elif output_path is not None and output_format == "json":
        print(f"  - Zusatz-JSON: {output_path.with_suffix('.json')}")
    elif output_path is not None and output_format == "yaml":
        print(f"  - Zusatz-YAML: {output_path.with_suffix('.yaml')}")
    if args.fill_excel:
        print(f"  - Excel: {Path(args.fill_excel)}")
    print("="*60)
    print("\nFertig!")


if __name__ == "__main__":
    main()
